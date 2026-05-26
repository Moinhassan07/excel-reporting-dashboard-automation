from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

file_path = "output/excel_reporting_dashboard.xlsx"
wb = load_workbook(file_path)

header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(color="FFFFFF", bold=True)
title_font = Font(bold=True, size=14)
center = Alignment(horizontal="center", vertical="center")

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)

dashboard_name = "City_Summary"
if dashboard_name in wb.sheetnames:
    ws = wb[dashboard_name]
    ws["A1"] = "EV Fleet Reporting Dashboard Summary"
    ws["A1"].font = title_font

wb.save(file_path)
print(f"Formatted workbook saved: {file_path}")